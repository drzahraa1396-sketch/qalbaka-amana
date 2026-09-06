import io, os, sqlite3
from datetime import date, datetime
from pathlib import Path
from copy import copy

import pandas as pd
import streamlit as st

try:
    import gspread
    from google.oauth2.service_account import Credentials
except Exception:
    gspread = None
    Credentials = None

from risk_tables import AGES, SBPS, CHOL, BMI, NONLAB_ROWS, LAB_ROWS

APP_DIR = Path(__file__).parent
DB_PATH = APP_DIR / "qalbak_amanah.db"

st.set_page_config(page_title="قلبك أمانة | مساعد الطبيب", page_icon="❤️", layout="wide")

HEADERS = {
    "Patients": [
        "national_id", "name", "dob", "sex", "family_file", "mobile",
        "governorate", "health_admin", "unit", "created_at"
    ],
    "Visits": [
        "visit_id", "visit_date", "national_id", "name", "family_file", "campaign_status",
        "age", "sex", "mobile", "governorate", "health_admin", "unit", "height_cm", "weight_kg", "bmi",
        "chol_mgdl", "ldl_mgdl", "sbp", "dbp", "diabetes", "hypertension", "diabetes_status",
        "hypertension_status", "ecg", "smoking", "family_history", "established_ascvd", "ckd",
        "tod", "multiple_rf", "pregnancy", "lactation", "risk_method", "risk_pct", "risk_color",
        "statin_needed", "statin_intensity", "statin_regimen", "bp_treatment", "dm_treatment",
        "statin_given", "aspirin_given", "health_education", "referral", "referral_reason",
        "referral_specialty", "referral_urgency", "next_followup", "doctor", "nurse", "created_at"
    ],
    "Followup": [
        "id", "national_id", "name", "scheduled_date", "visit_date", "status",
        "call_1_date", "call_1_status", "call_2_date", "call_2_status", "call_3_date", "call_3_status",
        "caller", "created_at"
    ],
    "Referrals": [
        "id", "referral_date", "national_id", "name", "family_file", "mobile", "reason", "specialty", "urgency",
        "followup_1_date", "followup_1_status", "followup_2_date", "followup_2_status", "followup_3_date",
        "followup_3_status", "feedback_treatment", "feedback_tests", "feedback_admission", "feedback_other",
        "followup_staff", "created_at"
    ],
}

RISK_HELP = {
    "أخضر": "خطورة أقل من 5%",
    "أصفر": "خطورة من 5% إلى أقل من 10%",
    "برتقالي": "خطورة من 10% إلى أقل من 20%",
    "أحمر": "خطورة من 20% إلى أقل من 30%",
    "أحمر داكن": "خطورة 30% أو أكثر",
}


def safe_num(v):
    try:
        if v in (None, "", "nan"): return None
        return float(v)
    except Exception:
        return None


def risk_color(p):
    if p is None:
        return "غير متاح"
    if p < 5: return "أخضر"
    if p < 10: return "أصفر"
    if p < 20: return "برتقالي"
    if p < 30: return "أحمر"
    return "أحمر داكن"


def age_group(age):
    if age < 40: return "18-40"
    if age <= 65: return "40-65"
    return ">65"


def bmi_group(bmi):
    if bmi < 30: return "<30"
    return ">30"


def sbp_group(sbp):
    if sbp < 120: return "<120"
    if sbp < 140: return "120-139"
    if sbp < 160: return "140-159"
    if sbp < 180: return "160-179"
    return "≥180"


def age_chart_group(age):
    for g in AGES:
        a, b = map(int, g.split("-"))
        if a <= age <= b:
            return g
    return None


def bmi_chart_group(bmi):
    if bmi < 20: return "<20"
    if bmi < 25: return "20-24"
    if bmi < 30: return "25-29"
    if bmi <= 35: return "30-35"
    return "≥35"


def chol_chart_group(chol_mmol):
    if chol_mmol < 4: return "<4"
    if chol_mmol < 5: return "4-4.9"
    if chol_mmol < 6: return "5-5.9"
    if chol_mmol < 7: return "6-6.9"
    return "≥7"


def calculate_risk(method, age, sex, smoker, sbp, bmi=None, chol_mgdl=None, diabetic=False):
    """Exact integer percentage from the supplied Egypt WHO 2019 chart; chart covers age 40-74."""
    if age < 40 or age > 74:
        return None, None
    ag = age_chart_group(age)
    sg = sbp_group(float(sbp))
    sex_i = 0 if sex == "ذكر" else 1
    smoker_i = 1 if smoker else 0
    if method == "BMI / بدون معمل":
        row = next(r for r in NONLAB_ROWS if r[0] == ag and r[1] == sg)[2]
        idx = (0 if sex_i == 0 else 10) + smoker_i * 5
        bgi = BMI.index(bmi_chart_group(float(bmi)))
        p = int(row[idx + bgi])
        return p, risk_color(p)
    if chol_mgdl is None:
        return None, None
    row = next(r for r in LAB_ROWS if r[0] == ag and r[1] == sg)
    vals = row[2] if not diabetic else row[3]
    idx = (0 if sex_i == 0 else 10) + smoker_i * 5
    ci = CHOL.index(chol_chart_group(float(chol_mgdl) / 38.7))
    p = int(vals[idx + ci])
    return p, risk_color(p)


def add_months(d, months):
    return (pd.Timestamp(d) + pd.DateOffset(months=months)).date()


def followup_options(visit_date, risk_pct):
    if risk_pct is None:
        return []
    if risk_pct < 5:
        return [("بعد 12 شهر", add_months(visit_date, 12))]
    if risk_pct < 10:
        return [("بعد 3 أشهر", add_months(visit_date, 3)), ("بعد 6 أشهر", add_months(visit_date, 6)), ("بعد 9 أشهر", add_months(visit_date, 9))]
    if risk_pct < 20:
        return [("بعد 3 أشهر", add_months(visit_date, 3)), ("بعد 6 أشهر", add_months(visit_date, 6))]
    return [("بعد 3 أشهر", add_months(visit_date, 3))]


def statin_recommendation(risk_pct, diabetic, tod, multiple_rf, ldl, chol, established_ascvd, pregnancy, lactation):
    if pregnancy or lactation:
        return "ممنوع حسب الدليل", "لا يوجد", "لا تستخدم الستاتين"
    if established_ascvd:
        return "نعم", "High-intensity", "Atorvastatin 40–80 mg أو Rosuvastatin 20–40 mg"
    if ldl is not None and ldl >= 190:
        return "نعم", "High-intensity", "Atorvastatin 40–80 mg أو Rosuvastatin 20–40 mg"
    if chol is not None and chol > 320:
        return "نعم", "High-intensity", "Atorvastatin 40–80 mg أو Rosuvastatin 20–40 mg"
    if risk_pct is not None and risk_pct > 30:
        return "نعم", "High-intensity", "Atorvastatin 40–80 mg أو Rosuvastatin 20–40 mg"
    if risk_pct is not None and risk_pct > 20:
        return "نعم", "Moderate-intensity", "Atorvastatin 20 mg أو Rosuvastatin 5–10 mg"
    if diabetic and (tod or multiple_rf):
        return "نعم", "High-intensity", "Atorvastatin 40–80 mg أو Rosuvastatin 20–40 mg"
    if diabetic:
        return "نعم", "Moderate-intensity", "Atorvastatin 20 mg أو Rosuvastatin 5–10 mg"
    return "لا", "لا يوجد", ""


def clinical_summary(age, bmi, diabetes, hypertension, smoking, family_history, risk_pct, ldl, established_ascvd, ckd):
    findings = []
    if established_ascvd:
        findings.append("ASCVD مثبتة")
    if ckd:
        findings.append("CKD مسجلة")
    if diabetes == "نعم":
        findings.append("سكري مسجل")
    if hypertension == "نعم":
        findings.append("ضغط مسجل")
    if bmi >= 30:
        findings.append("BMI ≥30")
    if smoking == "مدخن":
        findings.append("مدخن")
    if family_history == "يوجد":
        findings.append("تاريخ عائلي موجود")
    if ldl is not None and ldl >= 190:
        findings.append("LDL ≥190 mg/dL")
    if risk_pct is not None:
        findings.append(f"CVD Risk = {risk_pct}%")
    return findings


class Store:
    def __init__(self):
        self.gs = None
        if self.google_configured():
            try:
                self.gs = self._connect_gs()
            except Exception as e:
                st.warning(f"تعذر الاتصال بـ Google Sheets، سيتم استخدام التخزين المحلي: {e}")
        self.init_local()

    def google_configured(self):
        return gspread is not None and Credentials is not None and "gcp_service_account" in st.secrets and "google_sheet" in st.secrets

    def _connect_gs(self):
        cfg = dict(st.secrets["gcp_service_account"])
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_info(cfg, scopes=scopes)
        client = gspread.authorize(creds)
        ss = client.open(st.secrets["google_sheet"]["spreadsheet_name"])
        for title, headers in HEADERS.items():
            try:
                ws = ss.worksheet(title)
            except Exception:
                ws = ss.add_worksheet(title=title, rows=3000, cols=max(30, len(headers)))
                ws.append_row(headers)
                continue
            vals = ws.get_all_values()
            if not vals:
                ws.append_row(headers)
            elif vals[0] != headers:
                # Update the header row so the upgraded app can read older sheets safely.
                ws.update(f"A1:{chr(64 + min(len(headers), 26))}1", [headers[:26]]) if len(headers) <= 26 else ws.update("A1", [headers])
        return ss

    def init_local(self):
        con = sqlite3.connect(DB_PATH)
        for table, headers in HEADERS.items():
            cols = ",".join([f'"{h}" TEXT' for h in headers])
            con.execute(f'CREATE TABLE IF NOT EXISTS "{table}" ({cols})')
        con.commit()
        con.close()

    def df(self, table):
        headers = HEADERS[table]
        if self.gs:
            vals = self.gs.worksheet(table).get_all_values()
            if not vals:
                return pd.DataFrame(columns=headers)
            rows = vals[1:]
            normalized = []
            for row in rows:
                row = list(row) + [""] * max(0, len(headers) - len(row))
                normalized.append(row[:len(headers)])
            return pd.DataFrame(normalized, columns=headers)
        con = sqlite3.connect(DB_PATH)
        df = pd.read_sql_query(f'SELECT * FROM "{table}"', con)
        con.close()
        return df

    def append(self, table, row):
        headers = HEADERS[table]
        vals = [row.get(h, "") for h in headers]
        if self.gs:
            self.gs.worksheet(table).append_row(["" if v is None else str(v) for v in vals], value_input_option="USER_ENTERED")
        con = sqlite3.connect(DB_PATH)
        placeholders = ",".join(["?"] * len(headers))
        con.execute(f'INSERT INTO "{table}" ({",".join(headers)}) VALUES ({placeholders})', vals)
        con.commit()
        con.close()

    def exists_visit_month(self, national_id, year, month):
        df = self.df("Visits")
        if df.empty:
            return False
        x = pd.to_datetime(df["visit_date"], errors="coerce")
        return ((df["national_id"].astype(str) == str(national_id)) & (x.dt.year == year) & (x.dt.month == month)).any()

    def patient(self, national_id):
        df = self.df("Patients")
        if df.empty:
            return None
        m = df[df["national_id"].astype(str) == str(national_id)]
        return None if m.empty else m.iloc[-1].to_dict()

    def upsert_patient(self, row):
        old = self.patient(row["national_id"])
        if old is None:
            self.append("Patients", row)
        else:
            # Keep the master record fresh by appending a new version only when key data changed.
            changed = any(str(old.get(k, "")) != str(row.get(k, "")) for k in HEADERS["Patients"] if k != "created_at")
            if changed:
                self.append("Patients", row)


def make_excel(store):
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for table in HEADERS:
            store.df(table).to_excel(writer, index=False, sheet_name=table[:31])
    out.seek(0)
    return out


def monthly_report(store, year, month, governorate="", admin=""):
    visits = store.df("Visits")
    if visits.empty:
        return pd.DataFrame()
    d = pd.to_datetime(visits["visit_date"], errors="coerce")
    v = visits[(d.dt.year == year) & (d.dt.month == month)].copy()
    if governorate:
        v = v[v["governorate"].astype(str) == governorate]
    if admin:
        v = v[v["health_admin"].astype(str) == admin]
    if v.empty:
        return pd.DataFrame()
    rows = []
    for unit, g in v.groupby("unit", dropna=False):
        def n(mask): return int(mask.sum())
        age = pd.to_numeric(g.age, errors="coerce")
        bmi = pd.to_numeric(g.bmi, errors="coerce")
        risk = pd.to_numeric(g.risk_pct, errors="coerce")
        row = {
            "م": len(rows) + 1, "الادارة": admin or (g["health_admin"].dropna().iloc[0] if not g["health_admin"].dropna().empty else ""),
            "الوحدة /المركز": unit or "", "جديد": n(g.campaign_status.eq("جديد")), "متردد": n(g.campaign_status.eq("متردد")), "اجمالي": len(g),
            "18-40": n(age < 40), "40-65": n((age >= 40) & (age <= 65)), ">65": n(age > 65),
            "ذكر": n(g.sex.eq("ذكر")), "أنثى": n(g.sex.eq("أنثى")), "BMI <30": n(bmi < 30), "BMI >30": n(bmi >= 30),
            "سكر جديد": n((g.diabetes.eq("نعم")) & g.diabetes_status.eq("جديد")), "سكر متردد": n((g.diabetes.eq("نعم")) & g.diabetes_status.eq("متردد")),
            "ضغط جديد": n((g.hypertension.eq("نعم")) & g.hypertension_status.eq("جديد")), "ضغط متردد": n((g.hypertension.eq("نعم")) & g.hypertension_status.eq("متردد")),
            "سكر+ضغط جديد": n((g.diabetes.eq("نعم")) & (g.hypertension.eq("نعم")) & g.diabetes_status.eq("جديد") & g.hypertension_status.eq("جديد")),
            "سكر+ضغط متردد": n((g.diabetes.eq("نعم")) & (g.hypertension.eq("نعم")) & g.diabetes_status.eq("متردد") & g.hypertension_status.eq("متردد")),
            "رسم قلب جديد": n(g.ecg.eq("جديد")), "رسم قلب متابعة": n(g.ecg.eq("متابعة")), "طبيعي": n(g.ecg.eq("طبيعي")), "غير طبيعي": n(g.ecg.eq("غير طبيعي")),
            "مدخن": n(g.smoking.eq("مدخن")), "غير مدخن": n(g.smoking.eq("غير مدخن")),
            "تاريخ مرضي يوجد": n(g.family_history.eq("يوجد")), "تاريخ مرضي لا يوجد": n(g.family_history.eq("لا يوجد")),
            "<5": n(risk < 5), "5-10": n((risk >= 5) & (risk < 10)), "10-20": n((risk >= 10) & (risk < 20)), ">20": n(risk >= 20),
            "تثقيف صحي": n(g.health_education.eq("نعم")), "علاج ضغط": n(g.bp_treatment.eq("نعم")), "علاج سكر": n(g.dm_treatment.eq("نعم")),
            "ستاتين": n(g.statin_given.eq("نعم")), "اسبرين": n(g.aspirin_given.eq("نعم")), "احالة": n(g.referral.eq("نعم")),
        }
        rows.append(row)
    return pd.DataFrame(rows)


def report_xlsx(df, governorate, year, month):
    template = APP_DIR / "monthly_template.xlsx"
    if template.exists():
        from openpyxl import load_workbook
        wb = load_workbook(template)
        ws = wb["اسم المحافظه"]
        ws["G3"] = f"بيان قلبك امانة فى محافظة {governorate or ''} — شهر {month:02d}/{year}"
        data_start, total_row = 8, 23
        needed = len(df)
        if needed > 15:
            ws.insert_rows(total_row, needed - 15)
            total_row += needed - 15
            for r in range(23, total_row):
                for c in range(1, 38):
                    src, dst = ws.cell(8, c), ws.cell(r, c)
                    if src.has_style: dst._style = copy(src._style)
                    dst.number_format = src.number_format
                    dst.alignment = copy(src.alignment)
                ws.row_dimensions[r].height = ws.row_dimensions[8].height
        cols_map = {
            1:"م",2:"الادارة",3:"الوحدة /المركز",4:"جديد",5:"متردد",6:"اجمالي",7:"18-40",8:"40-65",9:">65",10:"ذكر",11:"أنثى",
            12:"BMI <30",13:"BMI >30",14:"سكر جديد",15:"سكر متردد",16:"ضغط جديد",17:"ضغط متردد",18:"سكر+ضغط جديد",19:"سكر+ضغط متردد",
            20:"رسم قلب جديد",21:"رسم قلب متابعة",22:"طبيعي",23:"غير طبيعي",24:"مدخن",25:"غير مدخن",26:"تاريخ مرضي يوجد",27:"تاريخ مرضي لا يوجد",
            28:"<5",29:"5-10",30:"10-20",31:">20",32:"تثقيف صحي",33:"علاج ضغط",34:"علاج سكر",35:"ستاتين",36:"اسبرين",37:"احالة"
        }
        for i in range(needed):
            r = data_start + i
            rec = df.iloc[i]
            for c, key in cols_map.items():
                val = rec.get(key, "")
                ws.cell(r, c).value = "" if pd.isna(val) else val
        for r in range(data_start + needed, total_row):
            for c in range(1, 38): ws.cell(r, c).value = None
        ws.cell(total_row, 1).value = "الاجمالى"
        for c in range(4, 38):
            ws.cell(total_row, c).value = sum(float(ws.cell(r, c).value or 0) for r in range(data_start, data_start + needed))
        out = io.BytesIO(); wb.save(out); out.seek(0); return out
    return io.BytesIO()


def set_patient_defaults(p):
    if not p: return
    mapping = {
        "national_id": p.get("national_id", ""), "name": p.get("name", ""), "family_file": p.get("family_file", ""),
        "mobile": p.get("mobile", ""), "governorate": p.get("governorate", ""), "health_admin": p.get("health_admin", ""),
        "unit": p.get("unit", ""), "sex": p.get("sex", "ذكر")
    }
    for k, v in mapping.items():
        st.session_state[k] = v or ""
    if p.get("dob"):
        try: st.session_state["dob"] = date.fromisoformat(str(p["dob"])[:10])
        except Exception: pass


store = Store()

# Header / navigation
st.sidebar.markdown("# ❤️ قلبك أمانة")
st.sidebar.caption("مساعد الطبيب — التسجيل والحساب والمتابعة والتقارير")
page = st.sidebar.radio("القائمة", [
    "🏠 الرئيسية", "➕ زيارة جديدة", "👤 ملف المريض", "📅 المتابعة والاستدعاء", "🚨 الإحالات", "📊 التقارير", "💾 تصدير البيانات", "⚙️ الإعدادات"
])

if not store.gs:
    st.sidebar.warning("التخزين الحالي محلي. للنشر الدائم على Streamlit Cloud فعّل Google Sheets.")
else:
    st.sidebar.success("الحفظ: Google Sheets")

# ---------------- HOME ----------------
if page == "🏠 الرئيسية":
    st.title("❤️ قلبك أمانة — مساعد الطبيب")
    today = date.today()
    visits = store.df("Visits")
    follow = store.df("Followup")
    refs = store.df("Referrals")
    if visits.empty:
        today_visits = pd.DataFrame()
        month_visits = pd.DataFrame()
    else:
        vd = pd.to_datetime(visits.visit_date, errors="coerce")
        today_visits = visits[vd.dt.date == today]
        month_visits = visits[(vd.dt.year == today.year) & (vd.dt.month == today.month)]
    if follow.empty:
        due_count = 0
    else:
        fd = pd.to_datetime(follow.scheduled_date, errors="coerce")
        due_count = int((fd.dt.date <= today).sum())
    ref_count = 0 if refs.empty else len(refs)
    risk_high = 0 if month_visits.empty else int(pd.to_numeric(month_visits.risk_pct, errors="coerce").ge(20).sum())

    a,b,c,d,e = st.columns(5)
    a.metric("زيارات اليوم", len(today_visits))
    b.metric("زيارات الشهر", len(month_visits))
    c.metric("متابعات مستحقة", due_count)
    d.metric("إحالات مسجلة", ref_count)
    e.metric("Risk ≥20% هذا الشهر", risk_high)

    st.divider()
    x,y = st.columns(2)
    with x:
        st.subheader("⚡ اختصارات الطبيب")
        if st.button("➕ تسجيل زيارة جديدة", use_container_width=True):
            st.session_state["go_to_visit"] = True
            st.rerun()
        st.info("سجّل المريض مرة واحدة؛ في الزيارات التالية استخدم الرقم القومي لاسترجاع بياناته.")
    with y:
        st.subheader("🚨 ما يحتاج انتباهًا")
        if due_count:
            st.warning(f"يوجد {due_count} موعد متابعة مستحق أو متأخر.")
        if risk_high:
            st.error(f"يوجد {risk_high} مريض/مرضى بخطورة ≥20% هذا الشهر.")
        if not due_count and not risk_high:
            st.success("لا توجد تنبيهات رئيسية حاليًا.")

    if not month_visits.empty:
        st.subheader("📊 ملخص الشهر الحالي")
        r = pd.to_numeric(month_visits.risk_pct, errors="coerce")
        s1,s2,s3,s4 = st.columns(4)
        s1.metric("جديد", int(month_visits.campaign_status.eq("جديد").sum()))
        s2.metric("متردد", int(month_visits.campaign_status.eq("متردد").sum()))
        s3.metric("مدخنون", int(month_visits.smoking.eq("مدخن").sum()))
        s4.metric("إحالات", int(month_visits.referral.eq("نعم").sum()))

# ---------------- NEW VISIT ----------------
elif page == "➕ زيارة جديدة":
    st.title("➕ تسجيل زيارة — شاشة الطبيب")
    st.caption("الفكرة: الطبيب يدخل البيانات مرة واحدة، والتطبيق يحسب ويجمع التقارير تلقائيًا.")

    st.subheader("1) المريض")
    c1,c2,c3 = st.columns([1.2,1.5,1.5])
    with c1:
        visit_date = st.date_input("تاريخ الزيارة", date.today(), key="visit_date")
        national_id = st.text_input("الرقم القومي *", max_chars=14, key="national_id")
        if st.button("🔎 بحث عن المريض", use_container_width=True):
            p = store.patient(national_id)
            if p:
                set_patient_defaults(p)
                st.session_state["found_patient"] = True
                st.success("تم العثور على المريض وتم تحميل بياناته.")
            else:
                st.session_state["found_patient"] = False
                st.info("المريض غير موجود — سيتم تسجيله كمريض جديد.")
    with c2:
        name = st.text_input("اسم المريض رباعي *", key="name")
        dob = st.date_input("تاريخ الميلاد", date(1980,1,1), min_value=date(1900,1,1), max_value=date.today(), key="dob")
        sex = st.selectbox("النوع", ["ذكر","أنثى"], key="sex")
        family_file = st.text_input("رقم الملف العائلي", key="family_file")
    with c3:
        mobile = st.text_input("رقم الموبايل", key="mobile")
        governorate = st.text_input("المحافظة", key="governorate")
        health_admin = st.text_input("الإدارة الصحية", key="health_admin")
        unit = st.text_input("الوحدة / المركز", key="unit")
        campaign_status = st.selectbox("حملة قلبك أمانة", ["جديد","متردد"])

    age = max(0, (visit_date - dob).days // 365)
    st.metric("العمر المحسوب", age)

    if store.exists_visit_month(national_id, visit_date.year, visit_date.month) if national_id else False:
        st.warning("⚠️ يوجد بالفعل تسجيل لهذا الرقم القومي في نفس الشهر. لن يسمح التطبيق بإنشاء زيارة ثانية.")

    st.subheader("2) القياسات")
    a,b,c,d,e = st.columns(5)
    with a: height = st.number_input("الطول (سم)", 80.0, 250.0, 165.0, key="height")
    with b: weight = st.number_input("الوزن (كجم)", 20.0, 300.0, 70.0, key="weight")
    bmi = round(weight / ((height/100)**2), 1)
    with c: st.metric("BMI", bmi)
    with d: sbp = st.number_input("SBP", 60, 260, 130, key="sbp")
    with e: dbp = st.number_input("DBP", 30, 160, 80, key="dbp")
    st.caption("يتم حفظ SBP وDBP معًا. جدول WHO المرفوع لحساب الخطر يستخدم SBP، وليس DBP.")

    st.subheader("3) الأمراض وعوامل الخطورة")
    a,b,c,d,e = st.columns(5)
    with a: diabetes = st.selectbox("السكري", ["لا","نعم"], key="diabetes")
    with b: hypertension = st.selectbox("الضغط", ["لا","نعم"], key="hypertension")
    with c: smoking = st.selectbox("التدخين", ["مدخن","غير مدخن"], key="smoking")
    with d: family_history = st.selectbox("تاريخ مرضي عائلي", ["يوجد","لا يوجد"], key="family_history")
    with e: ecg = st.selectbox("رسم القلب", ["غير مطلوب","جديد","متابعة","طبيعي","غير طبيعي"], key="ecg")
    a,b,c,d = st.columns(4)
    with a: diabetes_status = st.selectbox("حالة السكر", ["جديد","متردد","غير منطبق"], key="diabetes_status")
    with b: hypertension_status = st.selectbox("حالة الضغط", ["جديد","متردد","غير منطبق"], key="hypertension_status")
    with c: established_ascvd = st.checkbox("ASCVD مثبتة؟", key="ascvd")
    with d: ckd = st.checkbox("CKD؟", key="ckd")
    a,b,c,d = st.columns(4)
    with a: tod = st.checkbox("Target Organ Damage / TOD؟", key="tod")
    with b: multiple_rf = st.checkbox("Multiple additional risk factors؟", key="multiple_rf")
    with c: pregnancy = st.checkbox("حمل", key="pregnancy")
    with d: lactation = st.checkbox("رضاعة", key="lactation")

    st.subheader("4) التحاليل وتقييم الخطورة")
    method = st.radio("طريقة التقييم", ["معمل / Cholesterol","BMI / بدون معمل"], horizontal=True)
    chol = None; ldl = None
    if method == "معمل / Cholesterol":
        a,b = st.columns(2)
        with a: chol = st.number_input("Total Cholesterol (mg/dL)", 50.0, 600.0, 190.0, key="chol")
        with b: ldl = st.number_input("LDL (mg/dL)", 20.0, 400.0, 100.0, key="ldl")
    else:
        st.info("الطريقة غير المعملية تستخدم العمر + الجنس + التدخين + SBP + BMI طبقًا للجدول المرفوع.")

    risk_pct, risk_col = calculate_risk(method, age, sex, smoking == "مدخن", sbp, bmi, chol, diabetes == "نعم")
    if risk_pct is None:
        st.warning("تقييم WHO المرفوع متاح للأعمار 40–74 سنة فقط، أو يحتاج Total Cholesterol عند اختيار الطريقة المعملية.")
    else:
        st.success(f"❤️ CVD Risk خلال 10 سنوات: **{risk_pct}%** — {RISK_HELP.get(risk_col, risk_col)}")
        st.caption(f"طريقة الحساب: {method} | العمر {age} | SBP {sbp} | BMI {bmi}")

    st.subheader("5) مساعد القرار للطبيب")
    stat_needed, intensity, regimen = statin_recommendation(
        risk_pct, diabetes == "نعم", tod, multiple_rf, ldl, chol, established_ascvd, pregnancy, lactation
    )
    summary = clinical_summary(age, bmi, diabetes, hypertension, smoking, family_history, risk_pct, ldl, established_ascvd, ckd)
    if summary:
        st.write("**الانطباع/المشكلات المسجلة:** " + " • ".join(summary))
    a,b = st.columns(2)
    with a:
        st.markdown("### 💊 Statin")
        if stat_needed == "نعم": st.success(f"**مقترح:** {intensity}\n\n{regimen}")
        elif stat_needed == "ممنوع حسب الدليل": st.error("لا يستخدم الستاتين في الحمل/الرضاعة حسب الجايدلاين المرفوع.")
        else: st.info("لا توجد توصية Statin تلقائية من المعايير المدخلة.")
    with b:
        st.markdown("### 📅 المتابعة")
        opts = followup_options(visit_date, risk_pct)
        if opts:
            labels = [x[0] for x in opts]
            choice = st.selectbox("الفترة المقترحة", labels, key="fu_choice")
            next_fu = dict(opts)[choice]
            next_fu = st.date_input("التاريخ النهائي (يمكن للطبيب تعديله)", next_fu, key="next_fu")
        else:
            next_fu = None
            st.info("لا يوجد موعد WHO تلقائي لهذا العمر/النتيجة.")

    st.subheader("6) قرار الطبيب والإجراءات")
    a,b,c,d,e = st.columns(5)
    with a: health_education = st.selectbox("تثقيف صحي", ["نعم","لا"], key="health_education")
    with b: bp_treatment = st.selectbox("علاج ضغط", ["نعم","لا"], key="bp_treatment")
    with c: dm_treatment = st.selectbox("علاج سكر", ["نعم","لا"], key="dm_treatment")
    with d: statin_given = st.selectbox("ستاتين مصروف", ["نعم","لا"], key="statin_given")
    with e: aspirin_given = st.selectbox("أسبرين مصروف", ["نعم","لا"], key="aspirin_given")

    referral = st.selectbox("إحالة", ["لا","نعم"], key="referral")
    referral_reason = ""; referral_specialty = ""; referral_urgency = ""
    if referral == "نعم":
        a,b,c = st.columns(3)
        with a: referral_reason = st.text_input("سبب الإحالة")
        with b: referral_specialty = st.text_input("التخصص المحول إليه")
        with c: referral_urgency = st.selectbox("حالة الإحالة", ["عادية","طارئة"])
        st.warning("سيتم إنشاء متابعة إحالة بعد 3 أيام، مع إمكانية تسجيل موقف المريض والتغذية الراجعة لاحقًا.")

    a,b = st.columns(2)
    with a: doctor = st.text_input("اسم الطبيب ثلاثي", key="doctor")
    with b: nurse = st.text_input("اسم الممرضة ثلاثي", key="nurse")

    st.divider()
    st.subheader("✅ مراجعة قبل الحفظ")
    review = pd.DataFrame({"البند":["المريض","العمر","BMI","BP","Risk","Statin","المتابعة","الإحالة"],
                           "القيمة":[name or "-",age,bmi,f"{sbp}/{dbp}",f"{risk_pct}%" if risk_pct is not None else "غير متاح",f"{stat_needed} — {intensity}",next_fu.strftime("%d/%m/%Y") if next_fu else "-",referral]})
    st.dataframe(review, use_container_width=True, hide_index=True)

    if st.button("💾 اعتماد وحفظ الزيارة", type="primary", use_container_width=True):
        if not national_id or not name:
            st.error("الرقم القومي والاسم رباعي مطلوبان.")
        elif store.exists_visit_month(national_id, visit_date.year, visit_date.month):
            st.error("هذا المريض مسجل بالفعل في هذا الشهر — تم منع التكرار.")
        else:
            now = datetime.now().isoformat(timespec="seconds")
            store.upsert_patient({
                "national_id":national_id,"name":name,"dob":dob.isoformat(),"sex":sex,"family_file":family_file,"mobile":mobile,
                "governorate":governorate,"health_admin":health_admin,"unit":unit,"created_at":now
            })
            row = {
                "visit_id":f"{national_id}-{visit_date.isoformat()}","visit_date":visit_date.isoformat(),"national_id":national_id,"name":name,
                "family_file":family_file,"campaign_status":campaign_status,"age":age,"sex":sex,"mobile":mobile,"governorate":governorate,
                "health_admin":health_admin,"unit":unit,"height_cm":height,"weight_kg":weight,"bmi":bmi,"chol_mgdl":chol,"ldl_mgdl":ldl,
                "sbp":sbp,"dbp":dbp,"diabetes":diabetes,"hypertension":hypertension,"diabetes_status":diabetes_status,
                "hypertension_status":hypertension_status,"ecg":ecg,"smoking":smoking,"family_history":family_history,
                "established_ascvd":"نعم" if established_ascvd else "لا","ckd":"نعم" if ckd else "لا","tod":"نعم" if tod else "لا",
                "multiple_rf":"نعم" if multiple_rf else "لا","pregnancy":"نعم" if pregnancy else "لا","lactation":"نعم" if lactation else "لا",
                "risk_method":method,"risk_pct":risk_pct,"risk_color":risk_col,"statin_needed":stat_needed,"statin_intensity":intensity,
                "statin_regimen":regimen,"bp_treatment":bp_treatment,"dm_treatment":dm_treatment,"statin_given":statin_given,
                "aspirin_given":aspirin_given,"health_education":health_education,"referral":referral,"referral_reason":referral_reason,
                "referral_specialty":referral_specialty,"referral_urgency":referral_urgency,"next_followup":next_fu.isoformat() if next_fu else "",
                "doctor":doctor,"nurse":nurse,"created_at":now
            }
            store.append("Visits", row)
            if next_fu:
                store.append("Followup", {"id":f"{national_id}-{next_fu.isoformat()}","national_id":national_id,"name":name,
                    "scheduled_date":next_fu.isoformat(),"visit_date":visit_date.isoformat(),"status":"مجدول","created_at":now})
            if referral == "نعم":
                store.append("Referrals", {"id":f"{national_id}-{visit_date.isoformat()}","referral_date":visit_date.isoformat(),"national_id":national_id,
                    "name":name,"family_file":family_file,"mobile":mobile,"reason":referral_reason,"specialty":referral_specialty,
                    "urgency":referral_urgency,"followup_1_date":(pd.Timestamp(visit_date) + pd.Timedelta(days=3)).date().isoformat(),"created_at":now})
            st.success("تم اعتماد الزيارة. تم تحديث ملف المريض، المتابعة، الإحالة والإحصائيات تلقائيًا.")

# ---------------- PATIENT FILE ----------------
elif page == "👤 ملف المريض":
    st.title("👤 ملف المريض")
    nid = st.text_input("ابحث بالرقم القومي")
    if nid:
        p = store.patient(nid)
        v = store.df("Visits")
        if not v.empty: v = v[v.national_id.astype(str) == str(nid)].copy()
        if not p:
            st.warning("المريض غير موجود.")
        else:
            st.success(f"تم العثور على: {p.get('name','')}")
            a,b,c,d = st.columns(4)
            a.metric("الاسم", p.get("name", "-")); b.metric("النوع", p.get("sex", "-")); c.metric("الموبايل", p.get("mobile", "-")); d.metric("الملف العائلي", p.get("family_file", "-"))
            if not v.empty:
                v["visit_date"] = pd.to_datetime(v.visit_date, errors="coerce")
                latest = v.sort_values("visit_date").iloc[-1]
                st.subheader("آخر زيارة")
                a,b,c,d,e = st.columns(5)
                a.metric("التاريخ", latest.visit_date.strftime("%d/%m/%Y") if pd.notna(latest.visit_date) else "-")
                b.metric("BP", f"{latest.sbp}/{latest.dbp}")
                c.metric("BMI", latest.bmi)
                d.metric("Risk", f"{latest.risk_pct}%" if str(latest.risk_pct) not in ("", "nan") else "-")
                e.metric("Statin", latest.statin_intensity or "-")
                st.subheader("📜 تاريخ الزيارات")
                cols = ["visit_date","campaign_status","age","bmi","sbp","dbp","ldl_mgdl","risk_pct","risk_color","statin_given","referral","next_followup"]
                st.dataframe(v.sort_values("visit_date", ascending=False)[cols], use_container_width=True, hide_index=True)
            else:
                st.info("لا توجد زيارات مسجلة لهذا المريض.")

# ---------------- FOLLOWUP ----------------
elif page == "📅 المتابعة والاستدعاء":
    st.title("📅 المتابعة والاستدعاء")
    df = store.df("Followup")
    if df.empty:
        st.info("لا توجد سجلات متابعة.")
    else:
        df["scheduled_date"] = pd.to_datetime(df.scheduled_date, errors="coerce")
        today = date.today()
        due = df[df.scheduled_date.dt.date <= today].copy()
        upcoming = df[df.scheduled_date.dt.date > today].copy()
        a,b,c = st.columns(3)
        a.metric("مستحق/متأخر", len(due)); b.metric("القادم", len(upcoming)); c.metric("إجمالي المواعيد", len(df))
        st.subheader("🔴 المستحق والمتأخر")
        if due.empty: st.success("لا توجد متابعات مستحقة اليوم.")
        else: st.dataframe(due.sort_values("scheduled_date"), use_container_width=True, hide_index=True)
        st.subheader("🟢 المواعيد القادمة")
        st.dataframe(upcoming.sort_values("scheduled_date").head(100), use_container_width=True, hide_index=True)
        st.caption("الاستدعاء ليس زيارة جديدة. الزيارة الجديدة تُسجل فقط عند حضور المريض. يمكن لاحقًا تسجيل محاولات الاتصال وموقف المريض في نفس سجل المتابعة.")

# ---------------- REFERRALS ----------------
elif page == "🚨 الإحالات":
    st.title("🚨 الإحالات والتغذية الراجعة")
    df = store.df("Referrals")
    if df.empty:
        st.info("لا توجد إحالات.")
    else:
        rd = pd.to_datetime(df.referral_date, errors="coerce")
        st.metric("إجمالي الإحالات", len(df))
        st.dataframe(df.sort_values("referral_date", ascending=False), use_container_width=True, hide_index=True)
        st.caption("نموذج الإحالة يدعم سبب الإحالة، التخصص، درجة الاستعجال، والمتابعات والتغذية الراجعة.")

# ---------------- REPORTS ----------------
elif page == "📊 التقارير":
    st.title("📊 التقارير — بدون إعادة إدخال")
    today = date.today()
    a,b,c = st.columns(3)
    with a: year = st.number_input("السنة", 2020, 2100, today.year)
    with b: month = st.number_input("الشهر", 1, 12, today.month)
    with c: gov = st.text_input("المحافظة (اختياري)")
    admin = st.text_input("الإدارة الصحية (اختياري)")
    rep = monthly_report(store, int(year), int(month), gov, admin)
    if rep.empty:
        st.info("لا توجد بيانات لهذا الشهر/الفلاتر.")
    else:
        total = int(rep["اجمالي"].sum())
        new = int(rep["جديد"].sum()); follow = int(rep["متردد"].sum())
        a,b,c,d = st.columns(4)
        a.metric("الإجمالي", total); b.metric("جديد", new); c.metric("متردد", follow); d.metric("الوحدات", len(rep))
        st.dataframe(rep, use_container_width=True, hide_index=True)
        x = report_xlsx(rep, gov, int(year), int(month))
        st.download_button("⬇️ تحميل البيان الشهري بنفس قالب Excel", x, f"البيان_الشهري_قلبك_أمانة_{int(year)}_{int(month):02d}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------- EXPORT ----------------
elif page == "💾 تصدير البيانات":
    st.title("💾 تصدير البيانات")
    x = make_excel(store)
    st.download_button("⬇️ تحميل قاعدة البيانات كاملة Excel", x, "قلبك_أمانة_كل_البيانات.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    for t in HEADERS:
        st.write(f"**{t}**: {len(store.df(t))} سجل")

# ---------------- SETTINGS ----------------
elif page == "⚙️ الإعدادات":
    st.title("⚙️ الإعدادات")
    st.markdown("""
### Google Sheets على Streamlit Cloud
1. أنشئي Google Sheet باسم مثل `قلبك أمانة`.
2. أنشئي Service Account وفعّلي Google Sheets API وGoogle Drive API.
3. شاركي الـGoogle Sheet مع `client_email` الخاص بالـService Account كمحرر.
4. في Streamlit → Settings → Secrets ضعي `google_sheet.spreadsheet_name` وبيانات `gcp_service_account`.
5. عند التشغيل ستُنشأ تبويبات Patients / Visits / Followup / Referrals تلقائيًا.

**مهم:** النسخة الحالية مساعد قرار للطبيب وليست بديلًا عن الحكم السريري. التوصيات تُعرض حسب البيانات المدخلة والجايدلاين المرفوع، والطبيب يعتمد القرار النهائي.
    """)
    st.code('''[google_sheet]\nspreadsheet_name = "قلبك أمانة"\n\n[gcp_service_account]\ntype = "service_account"\nproject_id = "..."\nprivate_key_id = "..."\nprivate_key = "-----BEGIN PRIVATE KEY-----\\n...\\n-----END PRIVATE KEY-----\\n"\nclient_email = "..."\nclient_id = "..."\nauth_uri = "https://accounts.google.com/o/oauth2/auth"\ntoken_uri = "https://oauth2.googleapis.com/token"\nauth_provider_x509_cert_url = "https://www.googleapis.com/oauth2/v1/certs"\nclient_x509_cert_url = "..."''', language="toml")
